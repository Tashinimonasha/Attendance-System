from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, make_response, send_from_directory
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta, time
from PIL import Image, ImageEnhance
import pytesseract
import re
import os
import mysql.connector
import webbrowser
import threading
import hashlib
from functools import wraps
import logging
import random
from io import BytesIO

# Import configuration
from config import DB_CONFIG, ADMIN_CONFIG, COMPANY_MAPPING, COMPANY_DISPLAY_NAMES, TESSERACT_CMD, UPLOAD_FOLDER

# Set Tesseract path
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

def close_db_resources(cursor=None, connection=None):
    """Safely close database resources"""
    try:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
    except Exception as e:
        print(f"⚠️ Error closing database resources: {e}")

# Database initialization function
def initialize_database():
    """Initialize database and create tables if they don't exist"""
    try:
        # Connect to MySQL server (without selecting a database)
        connection = mysql.connector.connect(
            host=DB_CONFIG['host'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password']
        )
        
        if not connection:
            return False
            
        cursor = connection.cursor()
        
        # Create database if it doesn't exist
        cursor.execute("CREATE DATABASE IF NOT EXISTS AttendanceDB")
        cursor.execute("USE AttendanceDB")
        
        # Create AttendanceRecords table
        attendance_table = """
        CREATE TABLE IF NOT EXISTS `AttendanceRecords` (
            ID INT AUTO_INCREMENT PRIMARY KEY,
            NIC VARCHAR(20) NOT NULL,
            Date DATE NOT NULL,
            InTime TIME,
            OutTime TIME,
            Shift VARCHAR(50) DEFAULT 'Morning Shift',
            Status ENUM('Check in', 'Completed', 'Pending') DEFAULT 'Check in',
            Company ENUM('PUL', 'PCL', 'PPM', 'PDSL') DEFAULT 'PUL',
            FrontNICImage VARCHAR(255),
            BackNICImage VARCHAR(255),
            CreatedAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_nic (NIC),
            INDEX idx_date (Date),
            INDEX idx_shift (Shift),
            INDEX idx_company (Company),
            INDEX idx_status (Status)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
        cursor.execute(attendance_table)
        
        # Create admin_users table for admin authentication
        admin_table = """
        CREATE TABLE IF NOT EXISTS `admin_users` (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(64) NOT NULL,
            full_name VARCHAR(100),
            email VARCHAR(100),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
        cursor.execute(admin_table)
        
        # Create default admin user if it doesn't exist
        default_password = "printcare2025"
        hashed_password = hashlib.sha256(default_password.encode()).hexdigest()
        
        cursor.execute("""
            INSERT IGNORE INTO admin_users (username, password, full_name, email)
            VALUES (%s, %s, %s, %s)
        """, ('admin', hashed_password, 'System Administrator', 'admin@printcare.com'))
        
        connection.commit()
        close_db_resources(cursor, connection)
        
        print("✅ Database and tables initialized successfully")
        print("📋 Default admin credentials:")
        print("   Username: admin")
        print("   Password: printcare2025")
        
        return True
        
    except mysql.connector.Error as err:
        print(f"❌ Database initialization error: {err}")
        return False
    except Exception as e:
        print(f"❌ Unexpected error during database initialization: {e}")
        return False

# Initialize database on startup
initialize_database()

def populate_sample_data():
    """Populate database with comprehensive sample attendance data until today"""
    try:
        # Connect to database
        connection = get_db_connection()
        if not connection:
            print("❌ Failed to connect to database")
            return
            
        cursor = connection.cursor()
        
        # Check what's the latest date in the database
        cursor.execute("SELECT MAX(Date) FROM AttendanceRecords")
        latest_date_result = cursor.fetchone()[0]
        
        today = datetime.now().date()
        
        if latest_date_result:
            latest_date = latest_date_result
            if latest_date >= today:
                print(f"✅ Database is up to date (latest record: {latest_date})")
                close_db_resources(cursor, connection)
                return
            else:
                # Start from the day after the latest record
                start_date = datetime.combine(latest_date + timedelta(days=1), datetime.min.time())
                print(f"📅 Filling in missing attendance data from {start_date.date()} to {today}...")
        else:
            # No data exists, start from May 1, 2025
            start_date = datetime(2025, 5, 1)
            print(f"📅 Populating database with sample attendance data from {start_date.date()} to {today}...")
        
        # Define diverse workers with both old and new NIC formats
        workers = [
            # Old NIC format workers (ending with V/X)
            {'nic': '123456789V', 'company': 'PCL', 'name': 'Kasun Perera', 'department': 'Production'},
            {'nic': '987654321V', 'company': 'PPM', 'name': 'Saman Silva', 'department': 'Security'},
            {'nic': '456789123V', 'company': 'PDSL', 'name': 'Nimal Fernando', 'department': 'Administration'},
            {'nic': '789123456V', 'company': 'PUL', 'name': 'Ruwan Jayawardena', 'department': 'IT'},
            {'nic': '321654987V', 'company': 'PCL', 'name': 'Pradeep Kumara', 'department': 'Maintenance'},
            {'nic': '234567890V', 'company': 'PPM', 'name': 'Chaminda Rathnayake', 'department': 'Quality Control'},
            {'nic': '345678901V', 'company': 'PDSL', 'name': 'Buddhika Wijesiri', 'department': 'Production'},
            {'nic': '567890123V', 'company': 'PUL', 'name': 'Sampath Bandara', 'department': 'Security'},
            {'nic': '678901234V', 'company': 'PCL', 'name': 'Lakmal Dias', 'department': 'HR'},
            {'nic': '890123456V', 'company': 'PPM', 'name': 'Roshan Gunasekara', 'department': 'Finance'},
            {'nic': '159753468X', 'company': 'PDSL', 'name': 'Janaka Mendis', 'department': 'Logistics'},
            {'nic': '468135792X', 'company': 'PUL', 'name': 'Thilina Wickrama', 'department': 'Production'},
            {'nic': '357912468X', 'company': 'PCL', 'name': 'Upul Rajapaksa', 'department': 'Technical'},
            {'nic': '912345678X', 'company': 'PPM', 'name': 'Dinesh Abeysekara', 'department': 'Operations'},
            {'nic': '741963852X', 'company': 'PDSL', 'name': 'Mahinda Ranatunga', 'department': 'Cleaning'},
            
            # New NIC format workers (12 digits)
            {'nic': '200165432567', 'company': 'PUL', 'name': 'Amara Kodithuwakku', 'department': 'Production'},
            {'nic': '199587412369', 'company': 'PCL', 'name': 'Thushara Sanjeewa', 'department': 'IT'},
            {'nic': '198896745231', 'company': 'PPM', 'name': 'Kavinda Dissanayake', 'department': 'Security'},
            {'nic': '200278451236', 'company': 'PDSL', 'name': 'Nuwan Chathuranga', 'department': 'Administration'},
            {'nic': '199745123698', 'company': 'PUL', 'name': 'Gayan Madhushanka', 'department': 'Maintenance'},
            {'nic': '200012369874', 'company': 'PCL', 'name': 'Sachith Pathirana', 'department': 'Quality Control'},
            {'nic': '199512347896', 'company': 'PPM', 'name': 'Lakshan Perera', 'department': 'Production'},
            {'nic': '200145612378', 'company': 'PDSL', 'name': 'Chanaka Rathnayake', 'department': 'Security'},
            {'nic': '199887412365', 'company': 'PUL', 'name': 'Dhanuka Wijesinghe', 'department': 'HR'},
            {'nic': '200156789123', 'company': 'PCL', 'name': 'Tharaka Jayasundara', 'department': 'Finance'},
            {'nic': '199698741235', 'company': 'PPM', 'name': 'Nipuna Senanayake', 'department': 'Logistics'},
            {'nic': '200078945612', 'company': 'PDSL', 'name': 'Kasun Mallikarachchi', 'department': 'Production'},
            {'nic': '199523698741', 'company': 'PUL', 'name': 'Randika Wickramasinghe', 'department': 'Technical'},
            {'nic': '200189632574', 'company': 'PCL', 'name': 'Asanka Ranaweera', 'department': 'Operations'},
            {'nic': '199756123489', 'company': 'PPM', 'name': 'Dilan Senavirathne', 'department': 'Cleaning'},
            {'nic': '200098741256', 'company': 'PDSL', 'name': 'Chathura Karunaratne', 'department': 'Production'},
            {'nic': '199687452139', 'company': 'PUL', 'name': 'Buddhika Amarasinghe', 'department': 'Security'},
            {'nic': '200123987456', 'company': 'PCL', 'name': 'Lahiru Wijesekara', 'department': 'IT'},
            {'nic': '199512378965', 'company': 'PPM', 'name': 'Danushka Ekanayake', 'department': 'Administration'},
            {'nic': '200045612398', 'company': 'PDSL', 'name': 'Shehan Gunathilake', 'department': 'Maintenance'}
        ]
        
        # Define diverse shift schedules
        shift_schedules = {
            'Morning Shift': [
                ('08:00:00', '17:00:00'),
                ('08:30:00', '17:30:00'),
                ('09:00:00', '18:00:00')
            ],
            'Afternoon Shift': [
                ('13:00:00', '22:00:00'),
                ('14:00:00', '23:00:00'),
                ('15:00:00', '00:00:00')
            ],
            'Evening Shift': [
                ('18:00:00', '03:00:00'),
                ('19:00:00', '04:00:00'),
                ('20:00:00', '05:00:00')
            ],
            'Night Shift': [
                ('22:00:00', '07:00:00'),
                ('23:00:00', '08:00:00'),
                ('00:00:00', '09:00:00')
            ]  
             
        }
        
        # Generate data from start_date to today
        end_date = datetime.combine(today, datetime.min.time())
        current_date = start_date
        
        attendance_data = []
        
        while current_date <= end_date:
            # Skip some Sundays (70% chance to skip)
            if current_date.weekday() == 6 and random.random() < 0.7:
                current_date += timedelta(days=1)
                continue
                
            for worker in workers:
                # 88% attendance rate
                if random.random() < 0.88:
                    # Department-based shift assignment
                    department = worker['department']
                    if department == 'Production':
                        shift_options = ['Morning Shift', 'Afternoon Shift', 'Evening Shift', 'Production Shift']
                    elif department == 'Security':
                        shift_options = ['Security Shift', 'Night Shift', 'Evening Shift', 'Weekend Shift']
                    elif department in ['Administration', 'HR', 'Finance']:
                        shift_options = ['Morning Shift', 'Late Morning', 'Flexible Shift']
                    elif department == 'IT':
                        shift_options = ['Morning Shift', 'Flexible Shift', 'Late Morning', 'Afternoon Shift']
                    elif department in ['Maintenance', 'Technical']:
                        shift_options = ['Early Morning', 'Morning Shift', 'Afternoon Shift', 'Evening Shift']
                    elif department == 'Quality Control':
                        shift_options = ['Morning Shift', 'Afternoon Shift', 'Production Shift']
                    elif department == 'Logistics':
                        shift_options = ['Early Morning', 'Morning Shift', 'Afternoon Shift']
                    elif department == 'Operations':
                        shift_options = ['Morning Shift', 'Afternoon Shift', 'Evening Shift', 'Flexible Shift']
                    elif department == 'Cleaning':
                        shift_options = ['Early Morning', 'Evening Shift', 'Weekend Shift']
                    else:
                        shift_options = list(shift_schedules.keys())
                    
                    # Weekend workers get different shifts
                    if current_date.weekday() >= 5:  # Saturday or Sunday
                        if department in ['Security', 'Production', 'Maintenance']:
                            shift_options = ['Weekend Shift', 'Security Shift', 'Production Shift']
                        else:
                            # Reduced weekend staff
                            if random.random() < 0.3:  # Only 30% chance for non-essential departments
                                continue
                            shift_options = ['Weekend Shift', 'Flexible Shift']
                    
                    shift = random.choice(shift_options)
                    times = random.choice(shift_schedules[shift])
                    
                    # Add time variation (±20 minutes)
                    in_hour, in_min, _ = map(int, times[0].split(':'))
                    variation = random.randint(-20, 20)
                    in_min += variation
                    
                    # Handle minute overflow
                    if in_min >= 60:
                        in_hour += 1
                        in_min -= 60
                    elif in_min < 0:
                        in_hour -= 1
                        in_min += 60
                        
                    if in_hour >= 24:
                        in_hour -= 24
                    elif in_hour < 0:
                        in_hour += 24
                        
                    actual_in_time = f"{in_hour:02d}:{in_min:02d}:00"
                    
                    # Calculate out time (90% chance of checking out)
                    actual_out_time = None
                    if random.random() < 0.90:
                        # Department-based work duration
                        if department in ['Production', 'Security']:
                            work_duration = random.randint(480, 540)  # 8-9 hours
                        elif department in ['Administration', 'HR', 'Finance']:
                            work_duration = random.randint(450, 510)  # 7.5-8.5 hours
                        elif department == 'IT':
                            work_duration = random.randint(480, 600)  # 8-10 hours (flexible)
                        elif department in ['Maintenance', 'Technical']:
                            work_duration = random.randint(480, 660)  # 8-11 hours (overtime possible)
                        else:
                            work_duration = random.randint(480, 540)  # Standard 8-9 hours
                        
                        # Convert in_time to minutes
                        in_total_minutes = in_hour * 60 + in_min
                        out_total_minutes = in_total_minutes + work_duration
                        
                        # Handle day overflow
                        if out_total_minutes >= 1440:  # 24 hours = 1440 minutes
                            out_total_minutes -= 1440
                            
                        out_hour = out_total_minutes // 60
                        out_min = out_total_minutes % 60
                        
                        actual_out_time = f"{out_hour:02d}:{out_min:02d}:00"
                    
                    # Random status assignment
                    status_options = ['Check in', 'Completed', 'Pending']
                    status = random.choice(status_options)
                    
                    # Generate NIC image filenames
                    front_image = f"front_{worker['nic']}.jpg"
                    back_image = f"back_{worker['nic']}.jpg"
                    
                    attendance_data.append((
                        worker['nic'],
                        current_date.strftime('%Y-%m-%d'),
                        actual_in_time,
                        actual_out_time,
                        shift,
                        status,
                        worker['company'],
                        front_image,
                        back_image
                    ))
            
            current_date += timedelta(days=1)
        
        print(f"📊 Generated {len(attendance_data)} attendance records")
        
        # Insert all attendance data
        insert_query = """
        INSERT INTO `AttendanceRecords` (NIC, Date, InTime, OutTime, Shift, Status, Company, FrontNICImage, BackNICImage)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        cursor.executemany(insert_query, attendance_data)
        connection.commit()
        
        print(f"✅ Added {len(attendance_data)} attendance records to database")
        print(f"📅 Data spans from {start_date.date()} to {today}")
        print(f"👥 Covers {len(workers)} diverse workers")
        print(f"⏰ Includes {len(shift_schedules)} different shift types")
        
        close_db_resources(cursor, connection)
        
    except Exception as e:
        print(f"❌ Error populating sample data: {e}")

def delete_today_data():
    """Delete all attendance data for today (2025-07-12)"""
    try:
        # Connect to database
        connection = get_db_connection()
        if not connection:
            print("❌ Failed to connect to database")
            return
            
        cursor = connection.cursor()
        
        today = '2025-07-12'
        
        # Delete today's attendance records
        cursor.execute("DELETE FROM AttendanceRecords WHERE Date = %s", (today,))
        deleted_count = cursor.rowcount
        
        connection.commit()
        close_db_resources(cursor, connection)
        
        print(f"✅ Deleted {deleted_count} attendance records for {today}")
        
    except Exception as e:
        print(f"❌ Error deleting today's data: {e}")

# Delete today's data on startup - DISABLED to preserve user data
# delete_today_data()  # Commented out to prevent automatic deletion of attendance records

# Populate sample data on startup - DISABLED 
# populate_sample_data()  # Commented out to prevent automatic data generation

app = Flask(__name__)
app.secret_key = 'printcare_admin_secret_key_2025'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def get_shift_from_time(check_time):
    """
    Calculate shift based on check-in time
    
    Shift Schedule:
    - Morning Shift: 06:00 AM – 12:00 PM
    - Afternoon Shift: 12:00 PM – 06:00 PM  
    - Evening Shift: 06:00 PM – 10:00 PM
    - Night Shift: 10:00 PM – 06:00 AM (next day)
    """
    if isinstance(check_time, str):
        # Parse time string
        if ' ' in check_time:
            check_time = datetime.strptime(check_time, "%Y-%m-%d %H:%M:%S").time()
        else:
            check_time = datetime.strptime(check_time, "%H:%M:%S").time()
    elif isinstance(check_time, datetime):
        check_time = check_time.time()
    
    morning_start = time(6, 0)    # 06:00 AM
    afternoon_start = time(12, 0) # 12:00 PM
    evening_start = time(18, 0)   # 06:00 PM
    night_start = time(22, 0)     # 10:00 PM
    
    # Determine shift based on check-in time
    if morning_start <= check_time < afternoon_start:
        return "Morning Shift"
    elif afternoon_start <= check_time < evening_start:
        return "Afternoon Shift"
    elif evening_start <= check_time < night_start:
        return "Evening Shift"
    else:
        # Night shift: 10:00 PM - 06:00 AM (covers late night and early morning)
        return "Night Shift"

# Database connection utilities
def get_db_connection():
    """Get a fresh database connection with error handling"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG, autocommit=False)
        return connection
    except mysql.connector.Error as err:
        print(f"❌ Database connection error: {err}")
        return None
    except Exception as e:
        print(f"❌ Unexpected database error: {e}")
        return None

def get_simple_db_connection():
    """Get a simple database connection without selecting database"""
    try:
        config = DB_CONFIG.copy()
        config.pop('database', None)  # Remove database selection
        connection = mysql.connector.connect(**config)
        return connection
    except mysql.connector.Error as err:
        print(f"❌ Simple database connection error: {err}")
        return None

# Security utilities
def require_admin_auth(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            log_security_event('UNAUTHORIZED_ACCESS', f'Attempt to access {request.endpoint}')
            return jsonify({"error": "Unauthorized"}), 401
        
        # Check session timeout
        if 'admin_login_time' in session:
            login_time = datetime.fromisoformat(session['admin_login_time'])
            if datetime.now() - login_time > timedelta(minutes=ADMIN_CONFIG['session_timeout']):
                session.clear()
                log_security_event('SESSION_TIMEOUT', 'Admin session expired')
                return jsonify({"error": "Session expired"}), 401
        
        return f(*args, **kwargs)
    return decorated_function

def map_company_to_db(company):
    """Map company display name to database enum value"""
    return COMPANY_MAPPING.get(company, company if company in ['PUL', 'PCL', 'PPM', 'PDSL'] else 'PUL')

# Initialize database connection (using the centralized function)
try:
    db = get_db_connection()
    if db:
        cursor = db.cursor()
        print("✅ Database connected successfully")
    else:
        print("❌ Database connection failed")
        cursor = None
except Exception as e:
    print(f"❌ Database connection failed: {e}")
    db = None
    cursor = None

# Security tracking
security_log = []
failed_attempts = {}
locked_accounts = {}

def log_security_event(event_type, details="", ip_address="unknown"):
    """Log security events"""
    event = {
        'timestamp': datetime.now(),
        'type': event_type,
        'details': details,
        'ip_address': ip_address
    }
    security_log.append(event)
    print(f"🔒 SECURITY LOG: {event_type} - {details}")

def check_account_lockout(username):
    """Check if account is locked"""
    if username in locked_accounts:
        lockout_time = locked_accounts[username]
        if datetime.now() < lockout_time:
            return True
        else:
            del locked_accounts[username]
    return False

def verify_admin_credentials(username, password):
    """Enhanced admin verification with database lookup and security features"""
    # Check account lockout
    if check_account_lockout(username):
        log_security_event('LOCKOUT_ATTEMPT', f'Login attempt on locked account: {username}')
        return False
    
    try:
        connection = get_db_connection()
        if not connection:
            return False
            
        cursor = connection.cursor()
        
        # Hash the provided password
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute("""
            SELECT id, username, full_name FROM admin_users 
            WHERE username = %s AND password = %s
        """, (username, password_hash))
        
        user = cursor.fetchone()
        close_db_resources(cursor, connection)
        
        if user:
            log_security_event('LOGIN_SUCCESS', f'Admin {username} logged in successfully')
            # Clear failed attempts on successful login
            if username in failed_attempts:
                del failed_attempts[username]
            return True
        else:
            # Track failed attempts
            if username not in failed_attempts:
                failed_attempts[username] = []
            
            failed_attempts[username].append(datetime.now())
        
            # Clean old attempts (older than 1 hour)
            failed_attempts[username] = [
                attempt for attempt in failed_attempts[username]
                if datetime.now() - attempt < timedelta(hours=1)
            ]
            
            # Lock account if too many failures
            if len(failed_attempts[username]) >= ADMIN_CONFIG['max_failed_attempts']:
                lockout_until = datetime.now() + timedelta(minutes=ADMIN_CONFIG['lockout_duration'])
                locked_accounts[username] = lockout_until
                log_security_event('ACCOUNT_LOCKED', f'Account {username} locked due to failed attempts')
            
            log_security_event('LOGIN_FAILED', f'Failed login attempt for: {username}')
            return False
            
    except Exception as e:
        print(f"❌ Error during admin verification: {e}")
        return False

def extract_nic_from_text(text):
    """Lightning-fast NIC extraction from OCR text"""
    if not text:
        return None
        
    # Clean text - remove spaces and normalize (optimized)
    cleaned_text = ''.join(text.replace('\n', ' ').replace('\r', ' ').split()).upper()
    
    # Skip very short texts immediately
    if len(cleaned_text) < 9:
        return None
    
    # Only show cleaned text if it's potentially useful (reduced logging)
    if len(cleaned_text) >= 10:
        print(f"Cleaned text: {cleaned_text}")
    
    # Old NIC format: 9 digits + V/X (most common - try this first)
    old_nic_match = re.search(r'(\d{9})[VX]', cleaned_text)
    if old_nic_match:
        result = old_nic_match.group(0)
        print(f"⚡ OLD NIC FOUND: {result}")
        return result
    
    # New NIC format: 12 digits (try if old format not found)
    new_nic_match = re.search(r'\b(\d{12})\b', cleaned_text)
    if new_nic_match:
        result = new_nic_match.group(1)
        print(f"⚡ NEW NIC FOUND: {result}")
        return result
    
    # Alternative old format patterns (only if primary didn't work)
    for pattern in [r'(\d{2}\d{7})[VX]', r'(\d{3}\d{6})[VX]']:
        match = re.search(pattern, cleaned_text)
        if match:
            result = match.group(0)
            print(f"⚡ ALT OLD NIC: {result}")
            return result
    
    # Alternative new format patterns
    for pattern in [r'(\d{4}\d{8})', r'(\d{6}\d{6})']:
        match = re.search(pattern, cleaned_text)
        if match:
            result = match.group(1)
            print(f"⚡ ALT NEW NIC: {result}")
            return result
    
    return None

def extract_nic(img_path):
    """Ultra-fast NIC extraction optimized for maximum speed"""
    try:
        print(f"🔍 Processing image: {img_path}")
        
        # Load image once and reuse
        img = Image.open(img_path)
        
        # Convert to RGB if needed
        if img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Ultra-optimized OCR configs - only the most effective one first
        ultra_fast_configs = [
            '--psm 6 -c tessedit_char_whitelist=0123456789VXvx',  # Most effective for NICs
            '--psm 8 -c tessedit_char_whitelist=0123456789VXvx',  # Backup option
        ]
        
        # Super fast processing - only the most effective methods
        lightning_methods = [
            # Method 1: High contrast (most successful for NICs) - try this first
            lambda img: ImageEnhance.Contrast(img).enhance(2.0).convert('L'),
            # Method 2: Simple grayscale (fastest fallback)
            lambda img: img.convert('L'),
        ]
        
        # Try lightning-fast extraction - exit immediately when found
        for i, process_method in enumerate(lightning_methods):
            try:
                processed_img = process_method(img)
                
                for j, config in enumerate(ultra_fast_configs):
                    try:
                        # Fast OCR extraction
                        text = pytesseract.image_to_string(processed_img, config=config)
                        
                        # Immediate NIC extraction attempt
                        nic = extract_nic_from_text(text)
                        if nic:
                            print(f"⚡ LIGHTNING SUCCESS: {nic} (Method {i+1}.{j+1})")
                            return nic
                            
                    except Exception:
                        continue  # Skip failed attempts silently for speed
                        
            except Exception:
                continue  # Skip failed methods silently for speed
        
        print("⚡ Lightning extraction completed - no NIC found")
        return None
        
    except Exception as e:
        print(f"❌ Lightning OCR error: {e}")
        return None

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/scan', methods=['POST'])
def scan():
    """Optimized NIC scanning with ultra-fast processing"""
    if 'id_card' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400
    
    file = request.files['id_card']
    if file.filename == '':
        return jsonify({"success": False, "message": "No file selected"}), 400
    
    try:
        # Save uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Ultra-fast NIC extraction
        start_time = datetime.now()
        nic = extract_nic(filepath)
        end_time = datetime.now()
        
        extraction_time = (end_time - start_time).total_seconds()
        
        if nic:
            # Clean up file immediately after successful extraction
            try:
                os.remove(filepath)
            except:
                pass
                
            return jsonify({
                "success": True,
                "nic": nic,
                "extraction_time": round(extraction_time, 2),
                "message": "NIC extracted successfully"
            })
        else:
            # Clean up file even if extraction failed
            try:
                os.remove(filepath)
            except:
                pass
                
            return jsonify({
                "success": False,
                "message": "Could not extract NIC from image"
            }), 422
            
    except Exception as e:
        print(f"Scan error: {e}")
        return jsonify({
            "success": False,
            "message": f"Error processing image: {str(e)}"
        }), 500

@app.route('/capture_nic_images', methods=['POST'])
def capture_nic_images():
    """Capture NIC front and back images for attendance - IN ONLY"""
    try:
        data = request.get_json()
        action = data.get('action', '').upper()
        
        if action != 'IN':
            return jsonify({
                "success": False,
                "message": "NIC image capture is not required for OUT action - proceeding without images"
            }), 200  # Changed to 200 status to indicate this is expected behavior
        
        # Generate timestamp for unique filenames
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return jsonify({
            "success": True,
            "message": "Ready to capture NIC images for IN action",
            "timestamp": timestamp,
            "step": "front"
        })
        
    except Exception as e:
        print(f"NIC capture error: {e}")
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"
        }), 500

@app.route('/save_nic_image', methods=['POST'])
def save_nic_image():
    """Save captured NIC image (front or back) with NIC number as filename"""
    try:
        if 'nic_image' not in request.files:
            return jsonify({"success": False, "message": "No image uploaded"}), 400
        
        file = request.files['nic_image']
        image_type = request.form.get('image_type', 'front')  # 'front' or 'back'
        nic_number = request.form.get('nic_number', '')  # Get NIC number from form
        timestamp = request.form.get('timestamp', datetime.now().strftime("%Y%m%d_%H%M%S"))
        
        if file.filename == '':
            return jsonify({"success": False, "message": "No file selected"}), 400
            
        if not nic_number:
            return jsonify({"success": False, "message": "NIC number is required"}), 400
        
        # Create directories if they don't exist
        front_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'front')
        back_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'back')
        os.makedirs(front_dir, exist_ok=True)
        os.makedirs(back_dir, exist_ok=True)
        
        # Generate filename with NIC number
        if image_type == 'front':
            filename = f"{nic_number}_front.jpg"
            filepath = os.path.join(front_dir, filename)
        else:
            filename = f"{nic_number}_back.jpg"
            filepath = os.path.join(back_dir, filename)
        
        # Save the image (overwrite if exists)
        file.save(filepath)
        
        # Only extract NIC from front image for verification if needed
        extracted_nic = None
        if image_type == 'front':
            # Optionally extract NIC for verification, but don't require it
            try:
                extracted_nic = extract_nic(filepath)
            except Exception as e:
                print(f"NIC extraction from front image failed: {e}")
                # Continue without error since we already have NIC
        
        return jsonify({
            "success": True,
            "message": f"NIC {image_type} image saved as {filename}",
            "filename": filename,
            "filepath": filepath,
            "nic_number": extracted_nic if image_type == 'front' else nic_number,
            "image_type": image_type,
            "timestamp": timestamp
        })
        
    except Exception as e:
        print(f"Save NIC image error: {e}")
        return jsonify({
            "success": False,
            "message": f"Error saving image: {str(e)}"
        }), 500

@app.route('/record_attendance', methods=['POST'])
def record_attendance():
    """Record attendance to database with NIC images"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "status": "error",
                "message": "Invalid JSON data"
            }), 400
            
        # Safely handle None values and process strings
        action = data.get('action', '')
        nic = data.get('nic', '')
        company = data.get('company', '')
        front_image_path = data.get('front_image_path', '')
        back_image_path = data.get('back_image_path', '')
        
        # Convert None to empty string and process
        if action:
            action = str(action).strip().upper()
        if nic:
            nic = str(nic).strip()
        if company:
            company = str(company).strip()
        if front_image_path:
            front_image_path = str(front_image_path).strip()
        if back_image_path:
            back_image_path = str(back_image_path).strip()
        
        if not all([action, nic, company]):
            return jsonify({
                "status": "error",
                "message": "Missing required fields: action, nic, and company are required"
            }), 400
        
        # Use fresh database connection to avoid stale connections
        connection = get_db_connection()
        if not connection:
            return jsonify({
                "status": "error", 
                "message": "Database connection failed"
            }), 500
            
        cursor = connection.cursor()
        
        today = datetime.today().date()
        now = datetime.now().time()
        
        print(f"🔍 DEBUG: Processing {action} for NIC {nic} on {today} at {now}")
        
        # Map the company to database enum value using centralized function
        db_company = map_company_to_db(company)
        
        if action == 'IN':
            print(f"🔍 DEBUG: Processing IN for NIC {nic}, Company {company} -> {db_company}")
            
            # Check if this is a first-time user
            cursor.execute("""
                SELECT COUNT(*) FROM AttendanceRecords WHERE NIC = %s
            """, (nic,))
            
            total_records = cursor.fetchone()[0]
            is_first_time = total_records == 0
            
            # Calculate shift from time
            shift = get_shift_from_time(now)
            
            # Use provided image paths or empty strings if not provided
            front_img = front_image_path if front_image_path else ''
            back_img = back_image_path if back_image_path else ''
            
            print(f"🔍 DEBUG: About to insert - NIC: {nic}, Date: {today}, Time: {now}, Shift: {shift}, Company: {db_company}")
            
            cursor.execute(
                """INSERT INTO AttendanceRecords 
                   (NIC, Date, InTime, Shift, Status, Company, FrontNICImage, BackNICImage) 
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""", 
                (nic, today, now, shift, 'Check in', db_company, front_img, back_img)
            )
            connection.commit()
            
            print(f"✅ DEBUG: Successfully inserted attendance record for {nic}")
            
            # Prepare success message based on first-time status
            if is_first_time:
                message = f"🎉 Welcome to Printcare! Today is your first day. IN recorded for {nic} at Department {company}"
            else:
                message = f"✅ Welcome back! IN recorded for {nic} at Department {company}"
            
            return jsonify({
                "status": "success",
                "message": message,
                "time": now.strftime("%H:%M:%S"),
                "is_first_time": is_first_time
            })
            
        elif action == 'OUT':
            # 🏢 MULTI-COMPANY SUPPORT: Look for specific company's IN record without OUT time
            cursor.execute("""
                SELECT ID, InTime, OutTime, Status, Company FROM AttendanceRecords
                WHERE NIC = %s AND Date = %s AND Company = %s AND OutTime IS NULL
                ORDER BY InTime DESC
                LIMIT 1
            """, (nic, today, db_company))
            
            specific_company_record = cursor.fetchone()
            print(f"🔍 DEBUG: Looking for active {db_company} record for NIC {nic} on {today}")
            
            if specific_company_record:
                # Found an active IN record for the selected company
                record_id, in_time, out_time, status, record_company = specific_company_record
                print(f"✅ DEBUG: Found active {record_company} record ID {record_id} for OUT")
                
            else:
                # No active record for selected company, check if there are any records for this NIC today
                cursor.execute("""
                    SELECT ID, InTime, OutTime, Status, Company FROM AttendanceRecords
                    WHERE NIC = %s AND Date = %s
                    ORDER BY InTime DESC
                """, (nic, today))
                
                all_records = cursor.fetchall()
                
                if not all_records:
                    return jsonify({
                        "status": "warning",
                        "message": f"⚠️ No attendance records found for {nic} today. Please check IN first."
                    })
                
                # Check if there are any active IN records for other companies
                active_companies = []
                for record in all_records:
                    r_id, r_in_time, r_out_time, r_status, r_company = record
                    if r_out_time is None:  # Active IN record
                        active_companies.append(r_company)
                
                if active_companies:
                    # Show which companies have active IN records
                    company_list = ', '.join(active_companies)
                    return jsonify({
                        "status": "warning",
                        "message": f"⚠️ No active IN record found for {db_company}. Active companies for {nic}: {company_list}. Please select the correct company."
                    })
                else:
                    # All records are completed
                    return jsonify({
                        "status": "warning",
                        "message": f"⚠️ {nic} has already checked out from all companies today. No active IN records found."
                    })
            
            print(f"✅ DEBUG: Processing OUT for {db_company}. Updating record ID {record_id}")
            
            # Update with OUT time and status
            cursor.execute(
                "UPDATE AttendanceRecords SET OutTime = %s, Status = %s WHERE ID = %s", 
                (now, 'Completed', record_id)
            )
            connection.commit()
            print(f"✅ DEBUG: Record updated successfully")
            
            # Calculate work hours
            if isinstance(in_time, timedelta):
                in_seconds = int(in_time.total_seconds())
                in_time = datetime.min.time().replace(
                    hour=in_seconds // 3600,
                    minute=(in_seconds % 3600) // 60,
                    second=in_seconds % 60
                )
            
            in_datetime = datetime.combine(today, in_time)
            out_datetime = datetime.combine(today, now)
            work_duration = out_datetime - in_datetime
            
            hours = int(work_duration.total_seconds() // 3600)
            minutes = int((work_duration.total_seconds() % 3600) // 60)
            work_hours = f"{hours}h {minutes}m"
            
            return jsonify({
                "status": "success",
                "message": f"✅ OUT recorded for {nic} at Department {company}",
                "time": now.strftime("%H:%M:%S"),
                "work_hours": work_hours
            })
        
    except Exception as e:
        print(f"❌ Database error in record_attendance: {e}")
        print(f"❌ Error details - Action: {action if 'action' in locals() else 'unknown'}, NIC: {nic if 'nic' in locals() else 'unknown'}, Company: {company if 'company' in locals() else 'unknown'}")
        import traceback
        print(f"❌ Full traceback: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": f"Database error: {str(e)}"
        }), 500
    finally:
        # Always close the database connection
        close_db_resources(
            cursor if 'cursor' in locals() else None,
            connection if 'connection' in locals() else None
        )

@app.route('/admin')
def admin_login():
    """Secure admin login page"""
    if 'admin_logged_in' in session:
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html')

@app.route('/admin/login', methods=['POST'])
def admin_login_post():
    """Handle admin login with secure verification"""
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    
    if verify_admin_credentials(username, password):
        session['admin_logged_in'] = True
        session['admin_username'] = username
        session['admin_login_time'] = datetime.now().isoformat()
        flash('Login successful!', 'success')
        return redirect(url_for('admin_dashboard'))
    else:
        flash('Invalid username or password!', 'error')
        return redirect(url_for('admin_login'))

@app.route('/admin/logout')
def admin_logout():
    """Admin logout - redirect to main page with success message"""
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    session.pop('admin_login_time', None)
    flash('Logged out successfully!', 'success')
    return redirect('/')  # Redirect to main attendance page

@app.route('/admin/dashboard')
def admin_dashboard():
    """Modern admin dashboard with secure access"""
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    
    return render_template('modern_admin_dashboard.html')

@app.route('/admin/api/stats')
@require_admin_auth
def admin_stats():
    """API endpoint for admin statistics with date filtering"""
    try:
        cursor = db.cursor()
        
        # Get date parameter (default to today)
        selected_date = request.args.get('date', datetime.today().strftime('%Y-%m-%d'))
        
        # Get total unique employees for selected date (using AttendanceRecords table)
        cursor.execute("SELECT COUNT(DISTINCT NIC) FROM attendancerecords WHERE Date = %s", (selected_date,))
        total_employees = cursor.fetchone()[0] or 0
        
        # Get selected date's unique workers who attended  
        cursor.execute("SELECT COUNT(DISTINCT NIC) FROM attendancerecords WHERE Date = %s", (selected_date,))
        date_attendance = cursor.fetchone()[0] or 0
        
        # Get completed sessions (workers who have both IN and OUT times on selected date)
        cursor.execute("""
            SELECT COUNT(DISTINCT NIC) as completed_sessions
            FROM attendancerecords 
            WHERE Date = %s 
            AND InTime IS NOT NULL 
            AND OutTime IS NOT NULL
        """, (selected_date,))
        completed_sessions = cursor.fetchone()[0] or 0
        
        # Calculate total work hours for completed sessions
        cursor.execute("""
            SELECT SUM(TIME_TO_SEC(TIMEDIFF(OutTime, InTime))) / 3600 as total_hours
            FROM attendancerecords 
            WHERE Date = %s 
            AND InTime IS NOT NULL 
            AND OutTime IS NOT NULL
        """, (selected_date,))
        result = cursor.fetchone()
        total_hours = result[0] if result and result[0] else 0
        
        # Get currently active workers (those with IN but no OUT today)
        cursor.execute("""
            SELECT COUNT(DISTINCT NIC) as active_workers
            FROM attendancerecords 
            WHERE Date = %s 
            AND InTime IS NOT NULL 
            AND OutTime IS NULL
        """, (selected_date,))
        currently_in = cursor.fetchone()[0] or 0
        
        return jsonify({
            "selected_date": selected_date,
            "total_employees": total_employees,
            "date_attendance": date_attendance,
            "completed_sessions": completed_sessions,
            "date_hours": round(total_hours, 1),
            "currently_in": currently_in
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/admin/api/reports')
@require_admin_auth
def admin_reports():
    """API endpoint for admin reports"""
    report_type = request.args.get('type', 'daily')
    
    try:
        cursor = db.cursor()
        
        if report_type == 'daily':
            # Get last 7 days attendance
            cursor.execute("""
                SELECT 
                    DATE(time) as date, COUNT(DISTINCT nic) as attendance_count
                FROM attendance 
                WHERE DATE(time) >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                GROUP BY DATE(time) 
                ORDER BY date DESC
            """)
            
        elif report_type == 'weekly':
            # Get last 8 weeks attendance (by week)
            cursor.execute("""
                SELECT 
                    YEARWEEK(time) as week_year,
                    MIN(DATE(time)) as week_start,
                    COUNT(DISTINCT nic) as attendance_count
                FROM attendance 
                WHERE DATE(time) >= DATE_SUB(CURDATE(), INTERVAL 8 WEEK)
                GROUP BY YEARWEEK(time) 
                ORDER BY week_year DESC
            """)
            
        elif report_type == 'monthly':
            # Get last 12 months attendance (by month)
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(time, '%Y-%m') as month,
                    COUNT(DISTINCT nic) as attendance_count
                FROM attendance 
                WHERE DATE(time) >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(time, '%Y-%m') 
            """)
        else:
            return jsonify({"error": "Invalid report type"}), 400
        
        results = cursor.fetchall()
        
        # Format the data for charts
        data = []
        for row in results:
            if report_type == 'daily':
                data.append({
                    "date": str(row[0]),
                    "attendance": row[1],
                    "hours": round(row[2] if row[2] else 0, 2)
                })
            elif report_type == 'weekly':
                data.append({
                    "week": f"Week {row[0]}",
                    "date": str(row[1]),
                    "attendance": row[2],
                    "hours": round(row[3] if row[3] else 0, 2)
                })
            else:  # monthly
                data.append({
                    "month": row[0],
                    "date": str(row[1]),
                    "attendance": row[2],
                    "hours": round(row[3] if row[3] else 0, 2)
                })
        
        return jsonify({"data": data, "type": report_type})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/admin/api/recent-attendance')
@require_admin_auth
def get_recent_attendance():
    """Get recent attendance records for dashboard"""
    try:
        cursor = db.cursor()
        
        # Get recent attendance records (last 50 records)
        cursor.execute("""
            SELECT nic, company, DATE(time) as date, time, action
            FROM attendance 
            ORDER BY time DESC 
            LIMIT 50
        """)
        
        records = cursor.fetchall()
        
        recent_data = []
        for record in records:
            recent_data.append({
                'nic': record[0],
                'company': record[1] or 'N/A',
                'date': str(record[2]),
                'time': str(record[3]),
                'action': record[4]
            })
        
        return jsonify({
            "success": True,
            "records": recent_data
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/admin/api/search-attendance')
@require_admin_auth
def search_attendance():
    """Enhanced search attendance records with multiple filters"""
    try:
        cursor = db.cursor()
        
        # Get filter parameters
        nic_query = request.args.get('nic', '').strip()
        company_query = request.args.get('company', '').strip()
        date_query = request.args.get('date', '').strip()
        status_query = request.args.get('status', '').strip()
        
        # Build dynamic query for AttendanceRecords table
        base_query = """
            SELECT NIC, Company, Date, InTime, OutTime, Shift, Status
            FROM AttendanceRecords 
            WHERE 1=1
        """
        
        params = []
        
        if nic_query:
            base_query += " AND NIC LIKE %s"
            params.append(f'%{nic_query}%')
        
        if company_query:
            base_query += " AND Company LIKE %s"
            params.append(f'%{company_query}%')
        
        if date_query:
            base_query += " AND Date = %s"
            params.append(date_query)
        
        if status_query:
            if status_query == 'IN':
                base_query += " AND OutTime IS NULL AND InTime IS NOT NULL"
            elif status_query == 'OUT':
                base_query += " AND OutTime IS NOT NULL"
            else:
                base_query += " AND Status = %s"
                params.append(status_query)
        
        base_query += " ORDER BY Date DESC, InTime DESC LIMIT 100"
        
        cursor.execute(base_query, params)
        records = cursor.fetchall()
        
        search_data = []
        for record in records:
            search_data.append({
                'nic': record[0],
                'company': record[1] or 'N/A',
                'date': str(record[2]),
                'in_time': str(record[3]) if record[3] else 'N/A',
                'out_time': str(record[4]) if record[4] else 'N/A',
                'shift': record[5] or 'N/A',
                'status': record[6] or 'N/A'
            })
        
        return jsonify({
            "success": True,
            "records": search_data,
            "count": len(search_data)
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/admin/api/workers')
def get_workers():
    """API endpoint for Workers information - simplified for attendancerecords table"""
    if 'admin_logged_in' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        if not db or not db.is_connected():
            return jsonify({'error': 'Database connection lost'}), 500
            
        cursor = db.cursor()
        
        # Get all unique workers from AttendanceRecords table
        cursor.execute('''
            SELECT 
                NIC,
                Company,
                MAX(Date) as last_active,
                COUNT(*) as total_sessions,
                CASE 
                    WHEN MAX(Date) = CURDATE() AND 
                         EXISTS(SELECT 1 FROM AttendanceRecords ar2 
                                WHERE ar2.NIC = AttendanceRecords.NIC 
                                AND ar2.Date = CURDATE() 
                                AND ar2.InTime IS NOT NULL
                                AND ar2.OutTime IS NULL) THEN 'IN'
                    WHEN MAX(Date) = CURDATE() THEN 'OUT'
                    ELSE 'INACTIVE'
                END as current_status
            FROM AttendanceRecords 
            GROUP BY NIC, Company
            ORDER BY MAX(Date) DESC, NIC
        ''')
        
        records = cursor.fetchall()
        
        workers_data = []
        for record in records:
            workers_data.append({
                'nic': record[0] if record[0] else 'N/A',
                'name': f'Worker {record[0][-3:]}' if record[0] else 'Unknown',  # Generate name from NIC
                'company': record[1] if record[1] else 'N/A',
                'position': 'Employee',  # Default position
                'phone': 'N/A',  # No phone in simple structure
                'last_active': str(record[2]) if record[2] else 'Never',
                'total_sessions': record[3] if record[3] else 0,
                'current_status': record[4] if record[4] else 'INACTIVE'
            })
        
        print(f"✅ Workers API: Found {len(workers_data)} workers")
        
        return jsonify({
            'success': True,
            'workers': workers_data,
            'count': len(workers_data)
        })
        
    except Exception as e:
        print(f"❌ Error fetching workers: {e}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/admin/api/analytics')
@require_admin_auth
def get_analytics():
    """API endpoint for Analytics data with updated AttendanceRecords table"""
    try:
        filter_type = request.args.get('filter', 'daily')
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        # Get fresh connection
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        if filter_type == 'daily':
            # Daily hourly breakdown
            cursor.execute('''
                SELECT 
                    HOUR(InTime) as hour_period,
                    COUNT(*) as attendance_count,
                    COUNT(CASE WHEN OutTime IS NOT NULL THEN 1 END) as completed_count,
                    COUNT(CASE WHEN Status = 'Check in' THEN 1 END) as checkin_count,
                    COUNT(CASE WHEN Status = 'Completed' THEN 1 END) as completed_status_count
                FROM `AttendanceRecords`
                WHERE Date = %s AND InTime IS NOT NULL
                GROUP BY HOUR(InTime)
                ORDER BY hour_period
            ''', (date_param,))
            
        elif filter_type == 'weekly':
            # Weekly breakdown (7 days from selected date)
            cursor.execute('''
                SELECT 
                    DAYNAME(Date) as day_name,
                    COUNT(DISTINCT NIC) as attendance_count,
                    COUNT(*) as total_records,
                    COUNT(CASE WHEN OutTime IS NOT NULL THEN 1 END) as completed_count
                FROM `AttendanceRecords`
                WHERE Date BETWEEN DATE_SUB(%s, INTERVAL 6 DAY) AND %s
                GROUP BY Date, DAYNAME(Date)
                ORDER BY Date
            ''', (date_param, date_param))
            
        else:  # monthly
            # Monthly breakdown (6 months from selected date)
            cursor.execute('''
                SELECT 
                    MONTHNAME(Date) as month_name,
                    COUNT(DISTINCT NIC) as attendance_count,
                    COUNT(*) as total_records,
                    COUNT(CASE WHEN OutTime IS NOT NULL THEN 1 END) as completed_count
                FROM `AttendanceRecords`
                WHERE Date BETWEEN DATE_SUB(%s, INTERVAL 5 MONTH) AND %s
                GROUP BY YEAR(Date), MONTH(Date), MONTHNAME(Date)
                ORDER BY Date
            ''', (date_param, date_param))
        
        results = cursor.fetchall()
        
        # Format data for charts
        labels = []
        attendance = []
        completed = []
        
        for row in results:
            if filter_type == 'daily':
                labels.append(f"{row['hour_period']:02d}:00")
                attendance.append(row['attendance_count'])
                completed.append(row.get('completed_count', 0))
            elif filter_type == 'weekly':
                labels.append(row['day_name'][:3])
                attendance.append(row['attendance_count'])
                completed.append(row.get('completed_count', 0))
            else:
                labels.append(row['month_name'][:3])
                attendance.append(row['attendance_count'])
                completed.append(row.get('completed_count', 0))
        
        # Get department/company distribution
        cursor.execute('''
            SELECT 
                Company,
                COUNT(DISTINCT NIC) as worker_count,
                COUNT(*) as total_records
            FROM `AttendanceRecords`
            WHERE Date >= DATE_SUB(%s, INTERVAL 30 DAY)
            GROUP BY Company
            ORDER BY Company
        ''', (date_param,))
        
        dept_results = cursor.fetchall()
        
        # Format department data
        dept_labels = []
        dept_data = []
        
        for row in dept_results:
            dept_labels.append(row['Company'])
            dept_data.append(row['worker_count'])
        
        # Ensure all departments are represented
        all_companies = ['PCL', 'PUL', 'PPM', 'PDSL']
        dept_labels_complete = []
        dept_data_complete = []
        
        for company in all_companies:
            if company in dept_labels:
                idx = dept_labels.index(company)
                dept_labels_complete.append(company)
                dept_data_complete.append(dept_data[idx])
            else:
                dept_labels_complete.append(company)
                dept_data_complete.append(0)
        
        return jsonify({
            'success': True,
            'analytics': {
                'labels': labels,
                'attendance': attendance,
                'completed': completed,
                'departments': {
                    'labels': dept_labels_complete,
                    'data': dept_data_complete
                }
            }
        })
        
    except Exception as e:
        print(f"❌ Analytics error: {e}")
        return jsonify({'error': 'Database error'}), 500
    finally:
        close_db_resources(
            cursor if 'cursor' in locals() else None,
            connection if 'connection' in locals() else None
        )

# New comprehensive reporting endpoints
@app.route('/admin/api/reports/daily')
@require_admin_auth
def daily_report():
    """Generate daily attendance report"""
    try:
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        company = request.args.get('company', '')
        shift = request.args.get('shift', '')
        
        # Get fresh connection
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Build query with filters
        query = """
        SELECT 
            NIC, Date, InTime, OutTime, Shift, Status, Company
        FROM `AttendanceRecords`
        WHERE Date = %s
        """
        params = [date_param]
        
        if company:
            query += " AND Company = %s"
            params.append(company)
            
        if shift:
            query += " AND Shift = %s"
            params.append(shift)
            
        query += " ORDER BY InTime"
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # Calculate summary statistics
        total_records = len(records)
        checked_in = len([r for r in records if r['Status'] == 'Check in'])
        checked_out = len([r for r in records if r['Status'] == 'Completed'])
        
        # Convert time to string for JSON serialization
        for record in records:
            if record['InTime']:
                record['InTime'] = str(record['InTime'])
            if record['OutTime']:
                record['OutTime'] = str(record['OutTime'])
            if record['Date']:
                record['Date'] = str(record['Date'])
        
        return jsonify({
            'success': True,
            'date': date_param,
            'summary': {
                'total': total_records,
                'checked_in': checked_in,
                'checked_out': checked_out,
                'active': checked_in - checked_out if checked_in > checked_out else 0
            },
            'records': records
        })
        
    except Exception as e:
        print(f"❌ Daily report error: {e}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    finally:
        close_db_resources(
            cursor if 'cursor' in locals() else None,
            connection if 'connection' in locals() else None
        )

@app.route('/admin/api/reports/weekly')
@require_admin_auth
def weekly_report():
    """Generate weekly attendance report"""
    try:
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        company = request.args.get('company', '')
        
        # Get fresh connection
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Weekly summary by day
        query = """
        SELECT 
            Date,
            DAYNAME(Date) as DayName,
            COUNT(*) as TotalRecords,
            COUNT(DISTINCT NIC) as UniqueWorkers,
            COUNT(CASE WHEN Status = 'Check in' THEN 1 END) as CheckedIn,
            COUNT(CASE WHEN Status = 'Completed' THEN 1 END) as Completed,
            COUNT(CASE WHEN Status = 'Pending' THEN 1 END) as Pending
        FROM `AttendanceRecords`
        WHERE Date BETWEEN DATE_SUB(%s, INTERVAL 6 DAY) AND %s
        """
        params = [date_param, date_param]
        
        if company:
            query += " AND Company = %s"
            params.append(company)
            
        query += " GROUP BY Date ORDER BY Date"
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        
        # Convert date to string for JSON serialization
        for result in results:
            if result['Date']:
                result['Date'] = str(result['Date'])
        
        return jsonify({
            'success': True,
            'period': f"Week ending {date_param}",
            'data': results
        })
        
    except Exception as e:
        print(f"❌ Weekly report error: {e}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    finally:
        close_db_resources(
            cursor if 'cursor' in locals() else None,
            connection if 'connection' in locals() else None
        )

@app.route('/admin/api/reports/monthly')
@require_admin_auth
def monthly_report():
    """Generate monthly attendance report"""
    try:
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        company = request.args.get('company', '')
        
        # Get fresh connection
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Get the first day of the month
        date_obj = datetime.strptime(date_param, '%Y-%m-%d')
        first_day = date_obj.replace(day=1).strftime('%Y-%m-%d')
        last_day = (date_obj.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        last_day = last_day.strftime('%Y-%m-%d')
        
        # Monthly summary by company and shift
        query = """
        SELECT 
            Company,
            Shift,
            COUNT(*) as TotalRecords,
            COUNT(DISTINCT NIC) as UniqueWorkers,
            COUNT(CASE WHEN InTime IS NOT NULL THEN 1 END) as CheckedIn,
            COUNT(CASE WHEN OutTime IS NOT NULL THEN 1 END) as CheckedOut,
            AVG(CASE 
                WHEN InTime IS NOT NULL AND OutTime IS NOT NULL 
                THEN TIME_TO_SEC(TIMEDIFF(OutTime, InTime))/3600 
                ELSE NULL 
            END) as AvgWorkHours
        FROM `AttendanceRecords`
        WHERE Date BETWEEN %s AND %s
        """
        params = [first_day, last_day]
        
        if company:
            query += " AND Company = %s"
            params.append(company)
            
        query += " GROUP BY Company, Shift ORDER BY Company, Shift"
        
        cursor.execute(query, params)
        company_summary = cursor.fetchall()
        
        # Round average work hours
        for record in company_summary:
            if record['AvgWorkHours']:
                record['AvgWorkHours'] = round(record['AvgWorkHours'], 2)
        
        return jsonify({
            'success': True,
            'month': date_obj.strftime('%B %Y'),
            'period': f"{first_day} to {last_day}",
            'company_summary': company_summary
        })
        
    except Exception as e:
        print(f"❌ Monthly report error: {e}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    finally:
        close_db_resources(
            cursor if 'cursor' in locals() else None,
            connection if 'connection' in locals() else None
        )

@app.route('/admin/api/export/excel')
@require_admin_auth
def export_excel():
    """Export attendance data to Excel"""
    try:
        report_type = request.args.get('type', 'daily')
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        company = request.args.get('company', '')
        
        # Get fresh connection
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        if report_type == 'daily':
            query = """
            SELECT 
                NIC, Date, InTime, OutTime, Shift, Status, Company,
                CASE 
                    WHEN InTime IS NOT NULL AND OutTime IS NOT NULL 
                    THEN TIMEDIFF(OutTime, InTime)
                    ELSE NULL 
                END as WorkDuration
            FROM `AttendanceRecords`
            WHERE Date = %s
            """
            params = [date_param]
            filename = f"daily_report_{date_param}.xlsx"
            
        elif report_type == 'weekly':
            query = """
            SELECT 
                Date, NIC, InTime, OutTime, Shift, Status, Company
            FROM `AttendanceRecords`
            WHERE Date BETWEEN DATE_SUB(%s, INTERVAL 6 DAY) AND %s
            """
            params = [date_param, date_param]
            filename = f"weekly_report_{date_param}.xlsx"
            
        else:  # monthly
            date_obj = datetime.strptime(date_param, '%Y-%m-%d')
            first_day = date_obj.replace(day=1).strftime('%Y-%m-%d')
            last_day = (date_obj.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            last_day = last_day.strftime('%Y-%m-%d')
            
            query = """
            SELECT 
                Date, NIC, InTime, OutTime, Shift, Status, Company
            FROM `AttendanceRecords`
            WHERE Date BETWEEN %s AND %s
            """
            params = [first_day, last_day]
            filename = f"monthly_report_{date_obj.strftime('%Y-%m')}.xlsx"
            
        if company:
            query += " AND Company = %s"
            params.append(company)
            
        query += " ORDER BY Date, InTime"
        
        cursor.execute(query, params)
        records = cursor.fetchall()
        
        # Create Excel file
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        from io import BytesIO
        
        # Convert to DataFrame
        df = pd.DataFrame(records)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{report_type.capitalize()} Report"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style the header
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"❌ Excel export error: {e}")
        return jsonify({'error': f'Export error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals():
            connection.close()

@app.route('/admin/api/export-worker')
def export_worker():
    """API endpoint to export worker data as CSV"""
    if 'admin_logged_in' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    nic = request.args.get('nic')
    if not nic:
        return jsonify({'error': 'NIC required'}), 400
    
    try:
        cursor = db.cursor()
        cursor.execute('''
            SELECT * FROM attendancerecords 
            WHERE nic = %s 
            ORDER BY date DESC, in_time DESC
        ''', (nic,))
        
        records = cursor.fetchall()
        
        # Convert to CSV format
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=['nic', 'company', 'date', 'in_time', 'out_time', 'workhours', 'status'])
        writer.writeheader()
        
        for record in records:
            writer.writerow(record)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=worker_{nic}_data.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': 'Export failed'}), 500

@app.route('/admin/api/check-auth')
def check_auth():
    """API endpoint to check if admin is authenticated"""
    try:
        if 'admin_logged_in' not in session:
            return jsonify({"authenticated": False}), 401
        
        # Check session timeout
        if 'admin_login_time' in session:
            login_time = datetime.fromisoformat(session['admin_login_time'])
            if datetime.now() - login_time > timedelta(minutes=ADMIN_CONFIG['session_timeout']):
                session.clear()
                log_security_event('SESSION_TIMEOUT', 'Admin session expired via API check')
                return jsonify({"authenticated": False, "reason": "session_expired"}), 401
        
        return jsonify({
            "authenticated": True,
            "username": session.get('admin_username', 'admin'),
            "login_time": session.get('admin_login_time')
        })
        
    except Exception as e:
        print(f"❌ Error checking authentication: {e}")
        return jsonify({"authenticated": False, "error": str(e)}), 500

def open_browser():
    """Open browser after server starts"""
    threading.Timer(1.5, lambda: webbrowser.open('http://127.0.0.1:5001')).start()

@app.route('/check_nic_status', methods=['POST'])
def check_nic_status():
    """Check if NIC user is first-time or returning user"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "success": False,
                "message": "Invalid JSON data"
            }), 400
            
        nic = data.get('nic', '')
        action = data.get('action', '')
        
        # Safely handle None values and strip/uppercase
        if nic:
            nic = str(nic).strip()
        if action:
            action = str(action).strip().upper()
        
        if not nic or not action:
            return jsonify({
                "success": False,
                "message": "NIC and action are required"
            }), 400
        
        # For OUT actions, never need images
        if action == 'OUT':
            return jsonify({
                "success": True,
                "is_first_time": False,
                "needs_images": False,
                "message": "OUT action - no images needed"
            })
        
        # For IN actions, check if user exists in database
        connection = get_db_connection()
        if not connection:
            return jsonify({
                "success": False,
                "message": "Database connection failed"
            }), 500
            
        cursor = connection.cursor()
        
        # Check if this NIC has any previous records with images
        cursor.execute("""
            SELECT COUNT(*) as record_count,
                   COUNT(CASE WHEN FrontNICImage IS NOT NULL AND FrontNICImage != '' THEN 1 END) as with_front,
                   COUNT(CASE WHEN BackNICImage IS NOT NULL AND BackNICImage != '' THEN 1 END) as with_back
            FROM AttendanceRecords 
            WHERE NIC = %s
        """, (nic,))
        
        result = cursor.fetchone()
        record_count, front_count, back_count = result
        
        # Check if NIC images exist on disk (only for users with database records)
        front_image_exists = False
        back_image_exists = False
        
        # Only check for disk images if user has existing database records
        if record_count > 0:
            front_path = os.path.join(app.config['UPLOAD_FOLDER'], 'front', f"{nic}_front.jpg")
            back_path = os.path.join(app.config['UPLOAD_FOLDER'], 'back', f"{nic}_back.jpg")
            
            if os.path.exists(front_path):
                front_image_exists = True
            if os.path.exists(back_path):
                back_image_exists = True
        
        # Determine if images are needed
        is_first_time = record_count == 0
        
        # First-time users ALWAYS need to capture images
        if is_first_time:
            needs_images = True
            has_complete_images = False
        else:
            # For returning users, check if they have complete images in DB or on disk
            has_complete_images = (front_count > 0 and back_count > 0) or (front_image_exists and back_image_exists)
            needs_images = not has_complete_images
        
        print(f"🔍 NIC Status Check for {nic}:")
        print(f"   Records: {record_count}, Front: {front_count}, Back: {back_count}")
        print(f"   Front file exists: {front_image_exists}, Back file exists: {back_image_exists}")
        print(f"   Is first time: {is_first_time}, Needs images: {needs_images}")
        
        cursor.close()
        connection.close()
        
        return jsonify({
            "success": True,
            "is_first_time": is_first_time,
            "needs_images": needs_images,
            "has_front_image": front_image_exists or front_count > 0,
            "has_back_image": back_image_exists or back_count > 0,
            "total_records": record_count,
            "message": "First time user - images required" if needs_images else "Returning user - no images needed"
        })
        
    except Exception as e:
        print(f"❌ Error checking NIC status: {e}")
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"

        }), 500

@app.route('/get_nic_company', methods=['POST'])
def get_nic_company():
    """Get all active companies for an NIC to help with OUT validation"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "success": False,
                "message": "Invalid JSON data"
            }), 400
            
        nic = data.get('nic', '')
        action = data.get('action', '')
        
        # Safely handle None values
        if nic:
            nic = str(nic).strip()
        if action:
            action = str(action).strip().upper()
        
        if not nic:
            return jsonify({
                "success": False,
                "message": "NIC is required"
            }), 400
        
        # Only provide this information for OUT actions
        if action != 'OUT':
            return jsonify({
                "success": True,
                "message": "Company validation not required for IN actions"
            })
        
        connection = get_db_connection()
        if not connection:
            return jsonify({
                "success": False,
                "message": "Database connection failed"
            }), 500
            
        cursor = connection.cursor()
        today = datetime.today().date()
        
        # Look for all active IN records (no OUT time) for this NIC today
        cursor.execute("""
            SELECT Company, InTime FROM AttendanceRecords
            WHERE NIC = %s AND Date = %s AND OutTime IS NULL
            ORDER BY InTime DESC
        """, (nic, today))
        
        results = cursor.fetchall()
        
        if not results:
            return jsonify({
                "success": False,
                "message": f"No active IN records found for {nic} today. Please check IN first.",
                "has_active_record": False,
                "active_companies": []
            })
        
        # Convert company codes to display names
        company_display_names = {
            'PUL': 'PUL',
            'PCL': 'PCL', 
            'PPM': 'PPM',
            'PDSL': 'PDSL'
        }
        
        active_companies = []
        for company, in_time in results:
            active_companies.append({
                "company": company,
                "display_company": company_display_names.get(company, company),
                "in_time": in_time.strftime("%H:%M:%S") if in_time else None
            })
        
        return jsonify({
            "success": True,
            "message": f"Found {len(active_companies)} active IN record(s) for {nic}",
            "has_active_record": True,
            "active_companies": active_companies,
            "total_active": len(active_companies)
        })
        
    except Exception as e:
        print(f"Get NIC company error: {e}")
        return jsonify({
            "success": False,
            "message": f"Error: {str(e)}"
        }), 500
    finally:
        close_db_resources(
            cursor if 'cursor' in locals() else None,
            connection if 'connection' in locals() else None
        )

if __name__ == '__main__':
    print("🚀 ATTENDANCE SYSTEM STARTING...")
    print("📊 Simple NIC extraction")
    if db:
        print("💾 Database: ✅ Connected")
    else:
        print("💾 Database: ❌ Running in Demo Mode")
    
    open_browser()
    app.run(debug=True, host='0.0.0.0', port=5001)
